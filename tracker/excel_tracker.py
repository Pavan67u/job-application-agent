import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from agents.state import JobListing


HEADERS = [
    "Job ID", "Date Applied", "Job Title", "Company", "Location",
    "Source", "Match Score", "Status", "Resume Version",
    "Job URL", "Skills Required", "Notes", "Follow Up Date", "Response"
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL = PatternFill("solid", fgColor="EEF2FF")
SOURCE_COLORS = {
    "linkedin": "0A66C2",
    "naukri": "E94040",
    "indeed": "2164F3",
}
STATUS_COLORS = {
    "applied": "00B050",
    "pending": "FFC000",
    "failed": "FF0000",
    "skipped": "808080",
    "interview": "7030A0",
    "rejected": "C00000",
    "offer": "00B0F0",
}

COL_WIDTHS = [10, 14, 30, 25, 20, 12, 13, 12, 22, 50, 40, 30, 15, 20]


def _style_header(ws):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF")
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30


def _color_cell(cell, hex_color: str):
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def init_tracker(path: str) -> None:
    """Create tracker if not exists."""
    if os.path.exists(path):
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()

    # ── Main Sheet ──────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Dashboard Sheet ─────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Job Application Dashboard"
    dash["A1"].font = Font(bold=True, size=16, color="1F3864", name="Arial")
    dash["A3"] = "Total Applied"
    dash["B3"] = '=COUNTA(Applications!B:B)-1'
    dash["A4"] = "LinkedIn"
    dash["B4"] = '=COUNTIF(Applications!F:F,"linkedin")'
    dash["A5"] = "Naukri"
    dash["B5"] = '=COUNTIF(Applications!F:F,"naukri")'
    dash["A6"] = "Indeed"
    dash["B6"] = '=COUNTIF(Applications!F:F,"indeed")'
    dash["A8"] = "Interviews"
    dash["B8"] = '=COUNTIF(Applications!H:H,"interview")'
    dash["A9"] = "Offers"
    dash["B9"] = '=COUNTIF(Applications!H:H,"offer")'
    dash["A10"] = "Rejected"
    dash["B10"] = '=COUNTIF(Applications!H:H,"rejected")'
    dash["A12"] = "Avg Match Score"
    dash["B12"] = '=IFERROR(AVERAGEIF(Applications!G:G,">"&0),0)'

    for row in [3, 4, 5, 6, 8, 9, 10, 12]:
        dash[f"A{row}"].font = Font(bold=True, name="Arial")
        dash[f"B{row}"].font = Font(name="Arial", color="1F3864")
    dash.column_dimensions["A"].width = 20
    dash.column_dimensions["B"].width = 15

    wb.save(path)
    print(f"[Tracker] Created: {path}")


def log_jobs(jobs: list[JobListing], path: str) -> None:
    """Append applied jobs to tracker."""
    init_tracker(path)
    wb = load_workbook(path)
    ws = wb["Applications"]

    next_row = ws.max_row + 1

    for i, job in enumerate(jobs):
        row_num = next_row + i
        alt = row_num % 2 == 0

        values = [
            job.id,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            job.title,
            job.company,
            job.location,
            job.source,
            job.match_score,
            job.application_status,
            job.resume_version,
            job.url,
            ", ".join(job.skills_required[:8]),
            job.error_msg if job.error_msg else "",
            "",  # Follow up date — manual
            "",  # Response — manual
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [3, 11, 12]))

            if alt:
                cell.fill = ALT_FILL

        # Source badge color
        source_cell = ws.cell(row=row_num, column=6)
        src_color = SOURCE_COLORS.get(job.source.lower(), "555555")
        _color_cell(source_cell, src_color)

        # Status badge color
        status_cell = ws.cell(row=row_num, column=8)
        st_color = STATUS_COLORS.get(job.application_status.lower(), "808080")
        _color_cell(status_cell, st_color)

        # Match score color scale indicator
        score_cell = ws.cell(row=row_num, column=7)
        score_cell.number_format = '0.0"%"'
        score_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_num].height = 22

    # Apply conditional color scale to match score column (G)
    last_row = ws.max_row
    if last_row > 1:
        ws.conditional_formatting.add(
            f"G2:G{last_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FF0000",
                mid_type="num", mid_value=70, mid_color="FFFF00",
                end_type="num", end_value=100, end_color="00B050"
            )
        )

    wb.save(path)
    print(f"[Tracker] Logged {len(jobs)} jobs -> {path}")
